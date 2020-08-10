import openpyxl as xl
from openpyxl.chart import BarChart, Reference

print("Hello Tenzou!")
art = '''
        .. .':loddl;.   ....                .':oddc'.    .,,;;,'"'"'"'"'",,""...                                               
               ...      ....'"'"'"'"',;,..     ..,oo.    ...;;,,;::;;;;;;:;;".
                          ....',,,,;;;:ldoc,.. .,,',,.   ..:dxddxkdc::;;;;;;;,,..'..   ........
                                        .,lkx,  .':c:,..',,;;;cllcc:::::::::::::::;,,'',;;:::::;;,'..
                                          .cko'    .;;,,,,,,,,;::::;;::::::::::;,,;;:::;;;::;;:::::::;,..
                                          ...,'.  .';;;;:;,,;:::::,,;::::::::;,,;:::::::::;;,',::::::::::;,..
                                        ...      ..,;:;,,;::::::;,,:::::::::;,;:::::::::::::,.,;:::::::;::::;,..
                                      .,;:'      .',,,,;;::::::;,;:::::::::;;;;:::::::::::::,',::::;:::;;,;::::'
                                     .;;;,.      ..',;:::;;:::;;;:::::::::;,;,;:::::::::;;:;;,'';::;;::::;,,;;.
                                     .....        .';:::;;:::;;:clllllc::;;;,,coooddddolclllx0x;,::::;;::::;'.
                         ...            ..        .:::;,;lddddxkkxxxdoc::,""";ccccclcccclolo0NKl';::::;;:::::'                 
                  .',;::,'..        ..;lxdl;..  .;:lddl:oxxxololcc::::::;'.',:::::::::coc:;oXWXd,;:;::;:ol::::,.               
                 .okdodl,..   ...';:okOOdc:::;,,xx:ll;,,::,;;::::::::::;,.'',:::::::::oo::;dNNNx',:;:::;cddddl:,.              
                  .:;...... ..,:ldddlc;'.    .'kNx::,',::;,,;:::::::::;;,..';:::::::::xo:,,xNNNx',:,;::;,:::okxo,              
                    .....      .....         .lNWk:;,';::;',;::::::::;,;'.;;;:::::::;lOo:;,xNNNx;:;,;:::,;:;';loc.             
                       .                     ;KWWO:,'':::;'';::::::::'',''::;;;;;;,,'lkc;::kNNNxll;,;:::;;::,';::,.            
                          ......           .kWWWKl,',:::,.';:::::::;',;.'c:;::::::;,ld:,:lkNWXxdo;,;:::;,;:,',:::.            
                           .,,""'..........'oXWWWNx,.;:::,.';:::::::;',,';cc:cclc::;;xXo;:dKNWKkOo,,;:::;,,;;",;::"
                          .',,,,'',,''"'':kKXNNNWWK:.;:::,.';;::::::,.........'',;ldx0Nx:;xNNN0kx:'';:::;,',:,';::'            
                         .,,,,,,,,,'.',;dXWWWNXNWWWd';:::,.';;::::::;,,,.......,codOXNNX0k0NNN0Od:;,;;::,'',;,....             
                        ..'',,,,,,'',;l0NWWWWXKNWWWk,,:::,.',;::::::;,;;::,','.:kXXXNNNNNNNNNKx::coko,,;,'',;,.                
                          ....',,'';:dXNWWNWNK0XWNXx;,:::,.',;::::::;,;::dkOkdokXNNNNNNNNNNNNd...,lOd,;:;,',;,.                
                            .....,;:xXWNWWWNKO0XWNKx;,;:;,'',,;::::::,,::dXNWNNNNNNNNNNNNNNWKl',lON0c,;:;,',;'                 
                                .;cxNNWWNNX00O0XWK0k:';;;,'',,;::::::;,;:l0WNNWNNNNNNNNNNNNNKxoxKNNx,,::;,',,'                 
                                 'cdOXNNXKOxoox0K0OOc',,,,,.,,,:::::::,,;:dXWWNNNNNNNNNNNNNNNK0NNN0c,;:;;,',,.                 
                                 .,;:ldkkxl::;;x0OOOc.,,,,,'',,;::::::;,;::kNNNNNNNNNNNNNNNNNNXXNXd,,::;;,',,.                 
                                  .'',;;;;;;::;;:loo;'',,,,'',,;:::::::;,o0XNNNNNNNNNNNNNNNNNNNNNk;,;:;;,',,'                  
                                   .....','"'',''"'"''"'"',,''"',:;;:::;,lONNNNNNNNNNNNNNNNNNNNN0c,;;:;;.....
                                    ','""'"',,,,'"'"'"',,,'"'.'"cxddddddOKKNNNNNNNNNNNNNNNNNNNNKc,,;:::'
                                     .,,,''"'"''.'"','',,,,,,,,,l0XKXXXXXNNNNNNNNNNNNNXXNNNNNNKl,''..,'.
                                      .''"'"'"'"'','.''.',,,,,,,:kKKXXNNNNNNNNNNNNNNNNNNNNNNXOc,,,.                            
                                        .'"'',,,'"'',,'.',,,,,,,,dKKK0OKNNNNNNNNNNNNNNNNNWN0o;,,,,.                            
                                          ..'"'"',,,,,,'',,,,,,,,c0KK0OkO0XNNNNNNNNNNNNNN0l,,,,,,,.
                  .:ol:;,'..                  ......'....',,,,,,,;xKKKKK0OkO0XNNNNNWNNWKd. .,,,,,'.
                ,dO0OOOOkkxxkd:.                         .,,;;;,,,l0KKKKKKK0OOO0KNWNNNk,   .',,,,'
              ,x00OOOOOOO0kkXWWXkl'       .,::;..         ',;::;;,,cx0KKKKKKKK0Oo:ldxl.     .,,,,'
            .oKKOOOOOOOOO0kkXWWWNWXk;.   :xO00Okooc........,;:::;;,',lxKKKKKKKKO,           .,,,,'
           .kX0OOOOOOOOOO0kkXWWNNNNWNk, ;xkOOOOOOkKKx:,,,'"';:::::;,',,lOKKKKKKo. ......... .',,,'                             
          .kXOkkOOOOOOOOOOOkO000KXKKKKxcokddkOOO0kkXWXd:;,,',::::::;''"';xKKKK0c.'''''''',,'..,,,,.
         .xX0OxddxkOOOOOOOOOOkdccllloodkOOOxdkOO0Ox0NWNNO;..';::::::;'.'',o0KKO;.'.......',,,'"',,.     'lxxdl:,.              
         :XX0OO0OkxxxkOOOOOxo:;'....'"',:cldddk0O0kkXWWWWOccc;::::;;:;',,',dKXk,.....''...'',,'..,'   ,xXWWWWWWNKxc.
         .oXNKOkxkkOkxkOOdc;,,,,,,,''"'"'..',;cxOOOkOXWNWWOloc;:::;,;:;cdxolxKk,..';cll:,...','..',.'xXWNWWWNX0OO00l.
           ;ONX0OkxOO0Odc,,,,,,,,,,,,''"'..''.',:okOdd0XKK0oloc;:::;,::;cdl;cddc:'.',codo::cclloc;;,dNWWWNKOkkxkOOk,           
            .xNNNK0OOxc,,,,,,,,,,,,,,,,,,''"'"''..;oxk0KXNXOooo:::::,;::;;,.',;cdc.'.'cddddddddo:;::c0WXOkxkO0OOOOl            
             l0000OOo,,,,,,,,,,,,,,,,,,,,,,,''.';cdkOOOkkKNWNOdl:;::;,;::,''"'.':;'.'';oddddddo:;::::oOkkO0O00O0OOd,.':cc,
            'x0O0O0k:',,,,,,,,,,,,,,,,,,,,,,,''c0XK0OOOOkkOXNNklc:;::;,:::;,,,,''..',,,ldddddodo:::;;;lkkkkkkkkkkOO0KXWWWNd.   
     ':;...'oOOOO00x:,,,,,,,,,,,,,,,,,,,,,,,,,'cKWWNXXK00000KKXOooc:::;,:::;,,,,'""',,,:llooodKXd;::,,:xKXXXXKXXXNNWNNWWWWNx.  
   .lO0OOkkOOOOOkxxx:,,,,,,,,,,,,,,,,,,,,,,,,,,'lXWNWWWNNXKKXK0KXKkc:::;,:::;',,'',,,,'',coldKWWKl;:;,;ckNWWNNNWWNWWNWWWWWWNc  
  .dOOOxddddxxxxxkOOc,,,,,,,,,,,,,,,,,,,,,,,,,,''lOK000OO0XNNN0kKNXkl:::;,:::;','',,,,'';xXOxOXWWOc::;;;cONWNWWWWWWWWWWWWWWWk. 
  'k0O0OOOkkkOO0OOOOo,,,,,,,,,,,,,,,,,,,,,,,,,''..;okOOOOkONWWKOOKNKOo:::,;:::;'',,,,,'';kWNKxONWNk:::;,;ckXXKKK00OOOOOO0KXXO' 
   :OOOOOOOOOOOOOOO0x:,,,,,,,,,,,,,,,,,,,,,,''.....':lodddkNWWNNXO0KOxl;::;;:::;:ddc;'',;kWNWKdkXWNx:::;;;:dkkkkkxxkkkkkOOxc.  
    ;kOOO0OOOOOOOOOOOl,,,,,,,,,,,,,,,,,,,,'.......'....'',kWWWNWWNX0xO0o:::;,::::lKX0dc,;OWNWNkxk0XXd:::;,;:okOOOO00OOOOl,.    
     'oOOOOOkxxxxkkkkx:,,,,,,,,,,,,,,,,''.....''"'"''.''.:OKXXNWWWWNkkKKkc::;;;:::lKWWNOd0WWWN0K0xO0Ol:::;;;;lOOkxxxOOOc       
      .,loc;''ckOOOOOOd;,,,,,,,,,,,,,''"'"',,,''"'"',::clxOOOOO0KXNNKKXK0xc::;;;:::oKWWN0KWNWX0XWKkk0kl:::;,;:dKNK0kxkk,       
               ;kOOOOOxc,,,,,,,,,,,',,,,,,,,,,,,''.;xOO00OkxxxxkOk0NNNWWNKkc::;;;:::dXWX0NWWWK0NWWXOkOko::::;;:l0WWWNKKk,      
               .x00kdlodc,,,,,,,'"',,,,,,,,,,,,,,'.:kOOOOkddkOO0kkXWWWWWNNXxc::;,;:::kN0KWWWN0KWNWWNX0kxl;:::;;;lKWWWWWWKd:.  
'''
print(art)
names = ['Neite', 'Tenzou', 'Alisa', 'Mushashi', 'Kazuno']


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


numbers = [12, 23, 43, 2, 3, 2, 3, 42]
print(numbers.sort())
print(numbers)
